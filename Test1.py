__author__ = '569882'
import openpyxl
import time
## dd/mm/yyyy format
date = time.strftime("%Y%m%d")

#from openpyxl.cell import get_column_letter, column_index_from_string
# c = Sheet['B1']
#
# # for i in range(1,8,1):
# # print(i ,Sheet.cell(row=i , column=2).value)
# row_length = Sheet.max_row
# col_length = Sheet.max_column
# print(col_length)
# #import openpyxl
print("Writing selected result")
#wb1 = openpyxl.Workbook()  # creates a new workbook
wb = openpyxl.load_workbook('incidentd.xlsx')  # Load the incident xlsx
wb1 = openpyxl.load_workbook('sc_taskd.xlsx')
wb2 = openpyxl.load_workbook('Lab Informatics Aging Report - 20171115.xlsx') # Load the DSR report
Incident = wb2.get_sheet_by_name('Incident')  # Grab the particular sheet in DSR and named it wb3
Task = wb2.get_sheet_by_name('TASK')  # Grab the particular sheet in DSR and named it wb3
rl_wb2 = Incident.max_row  # Maximum rows in DSR
cl_wb2 = Incident.max_column #Maximum column in DSR
# print(x)
# for i in range(2, row_length_inc_wb3):
#     for j in range(1, col_length_inc_wb3):
#         print(i, Incident.cell(row=i, column=j).value)
#row_lengthwb3 =
inc = wb.get_sheet_by_name('Page 1')# pulls the sheet with name as passed in arguments FOR Incidents
tsk = wb1.get_sheet_by_name('Page 1') # pulls the sheet with name as passed in arguments FOR Tasks
rl_wb = inc.max_row
cl_wb = inc.max_column
rl_wb1 = tsk.max_row
cl_wb1 = tsk.max_column
# s = wb1.active
# s.title ='Main'
#wb1.create_sheet("Tasks")
#sheet = wb1.get_active_sheet()
#sheet.title = 'Incidents'
for i in range(2, rl_wb+1):
    for j in range(1, 10):
        x = inc.cell(row=i, column=j).value
        Incident.cell(row=i, column=j).value = x
        y = tsk.cell(row=i, column=j).value
        Task.cell(row=i, column=j).value = y

for i in range(2, rl_wb+1):
        n = inc.cell(row=i, column=12).value
        Incident.cell(row=i, column=13).value = n
        m = tsk.cell(row=i, column=12).value
        Task.cell(row=i, column=13).value = m

Ag = wb2.get_sheet_by_name('Aging')
# sheet['A1'] = 'NEERAJ'
Incident['D1'] = 'Application'
name = 'Lab Informatics Aging Report - '+date+'.xlsx'
wb2.save(name)
print("Done")
# x = wb2.get_sheet_names()
